#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GP_ENG_2025.pdf'den ürün bilgilerini ve PriceList_2025_GBP.pdf'den fiyatları çıkarır
Excel formatına dönüştürür - Geliştirilmiş versiyon
"""

import pdfplumber
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Ürün kodu -> İngilizce isim -> Türkçe isim eşleştirmesi
product_names = {
    "32AUR": ("spare tip for automatic centre punch", "Otomatik zımba yedek uç"),
    "33LT": ("letter punches", "Harf zımba seti"),
    "33NR": ("number punches", "Rakam zımba seti"),
    "34": ("flat chisels", "Düz keskiler"),
    "35": ("flat chisels, ribbed type", "Düz keskiler, nervürlü tip"),
    "35PM": ("flat chisels, ribbed type with hand guards", "Düz keskiler, nervürlü tip, el korumalı"),
    "35PMR": ("spare hand guards", "Yedek el korumaları"),
    "36": ("cape chisels", "Sivri uçlu keskiler"),
    "37": ("cape chisels, ribbed type", "Sivri uçlu keskiler, nervürlü tip"),
    "37PM": ("cape chisel, ribbed type with hand guard", "Sivri uçlu keski, nervürlü tip, el korumalı"),
    "37PMR": ("spare hand guard", "Yedek el koruması"),
    "38/B6N": ("chisel set", "Keski seti"),
    "38/SP6": ("spare chisel tips", "Yedek keski uçları"),
    "42": ("combination wrenches", "Kombinasyon anahtarları"),
    "42AS": ("combination wrenches, open and offset ring ends", "Kombinasyon anahtarları, açık ve ofset halka uçlu"),
    "42MP": ("combination wrenches, open and offset ring ends, bright chrome-plated", "Kombinasyon anahtarları, parlak krom kaplı"),
    "42LMP": ("combination wrenches, long series, bright chrome-plated", "Kombinasyon anahtarları, uzun seri, parlak krom kaplı"),
    "42SLIM": ("combination wrenches with thin open ends", "İnce açık uçlu kombinasyon anahtarları"),
    "45": ("combination wrenches, heavy series", "Kombinasyon anahtarları, ağır seri"),
    "52": ("single open end wrenches", "Tek açık uçlu anahtarlar"),
    "53": ("single open end wrenches", "Tek açık uçlu anahtarlar"),
    "55": ("double open end wrenches", "Çift açık uçlu anahtarlar"),
    "55AS": ("double open end wrenches", "Çift açık uçlu anahtarlar"),
    "58": ("open end slogging wrenches", "Açık uçlu balyoz anahtarları"),
    "73": ("small double open end wrenches", "Küçük çift açık uçlu anahtarlar"),
    "78": ("ring slogging wrenches", "Halka balyoz anahtarları"),
    "78AS": ("ring slogging wrenches", "Halka balyoz anahtarları"),
    "80": ("double swivel end socket wrenches", "Çift mafsallı lokma anahtarları"),
    "83": ("half-moon ring wrenches", "Yarım ay halka anahtarları"),
    "83AS": ("half-moon ring wrenches", "Yarım ay halka anahtarları"),
    "88": ("double ended flat ring wrenches, extra-long series", "Çift uçlu düz halka anahtarları, ekstra uzun seri"),
    "90": ("double ended deep offset ring wrenches", "Çift uçlu derin ofset halka anahtarları"),
    "90AS": ("double ended deep offset ring wrenches", "Çift uçlu derin ofset halka anahtarları"),
    "91": ("heavy duty offset ring wrenches", "Ağır hizmet ofset halka anahtarları"),
    "92": ("tubes for item 91", "91 için borular"),
    "93": ("double-ended offset ring wrench for scaffolding bolts", "İskele cıvataları için çift uçlu ofset halka anahtar"),
    "93C": ("ratcheting double-ended offset ring wrenches for scaffolding bolts", "İskele cıvataları için cırcırlı çift uçlu ofset halka anahtarlar"),
    "94": ("flare nut open ring wrenches", "Fren borusu anahtarları"),
    "95": ("double ended flat ring wrenches", "Çift uçlu düz halka anahtarları"),
    "95FTX": ("double-ended straight wrenches for Torx head screws", "Torx başlı vidalar için çift uçlu düz anahtarlar"),
    "96": ("offset hexagon key wrenches, chrome-plated", "Ofset altıgen anahtarlar, krom kaplı"),
    "96N": ("offset hexagon key wrenches, burnished", "Ofset altıgen anahtarlar, parlatılmış"),
    "96AS": ("offset hexagon key wrenches, burnished", "Ofset altıgen anahtarlar, parlatılmış"),
    "96LC": ("offset hexagon key wrenches, long series, chrome-plated", "Ofset altıgen anahtarlar, uzun seri, krom kaplı"),
    "96L": ("offset hexagon key wrenches, long series, burnished", "Ofset altıgen anahtarlar, uzun seri, parlatılmış"),
    "96BP": ("ball head offset hexagon key wrenches, burnished", "Bilyalı uç ofset altıgen anahtarlar, parlatılmış"),
    "96BP-CL": ("ball head offset hexagon key wrenches, chrome-plated, coloured", "Bilyalı uç ofset altıgen anahtarlar, krom kaplı, renkli"),
    "96BPC": ("ball head offset hexagon key wrenches, chrome-plated", "Bilyalı uç ofset altıgen anahtarlar, krom kaplı"),
    "96LBP": ("ball head offset hexagon key wrenches, extra-long model", "Bilyalı uç ofset altıgen anahtarlar, ekstra uzun model"),
    "96BP-HO": ("ball head offset hexagon key wrenches with screw holding system", "Vida tutma sistemli bilyalı uç ofset altıgen anahtarlar"),
    "96T": ("offset hexagon key wrenches with high torque handles", "Yüksek torklu saplarla ofset altıgen anahtarlar"),
    "96T/AS": ("offset hexagon key wrenches with high torque handles", "Yüksek torklu saplarla ofset altıgen anahtarlar"),
    "96TBP": ("ball head offset hexagon key wrenches with high torque handles", "Yüksek torklu saplarla bilyalı uç ofset altıgen anahtarlar"),
    "96BPA": ("ball head offset hexagon key wrenches, 110°, extra-short side model", "Bilyalı uç ofset altıgen anahtarlar, 110°, ekstra kısa yan model"),
    "97TX": ("offset key wrenches for Torx head screws", "Torx başlı vidalar için ofset anahtarlar"),
    "97BTX": ("ball head offset key wrenches for Torx head screws", "Torx başlı vidalar için bilyalı uç ofset anahtarlar"),
    "97BTXL": ("ball head offset key wrenches, long model, for Torx head screws", "Torx başlı vidalar için bilyalı uç ofset anahtarlar, uzun model"),
    "97RTX": ("offset key wrenches for Tamper Resistant Torx head screws", "Güvenlikli Torx başlı vidalar için ofset anahtarlar"),
    "97BRTXL": ("ball head offset key wrenches, long model, for Tamper Resistant Torx head screws", "Güvenlikli Torx başlı vidalar için bilyalı uç ofset anahtarlar, uzun model"),
    "97TTX": ("offset key wrenches with handles for Torx head screws", "Torx başlı vidalar için saplı ofset anahtarlar"),
    "98XZN": ("offset key wrenches with XZN profile", "XZN profilli ofset anahtarlar"),
    "99": ("hook wrenches with square noses for ring nuts", "Segman somunları için kare uçlu kanca anahtarlar"),
    "99SQ": ("hook wrenches with square noses for ring nuts", "Segman somunları için kare uçlu kanca anahtarlar"),
    "99ST": ("hook wrenches with round noses for ring nuts", "Segman somunları için yuvarlak uçlu kanca anahtarlar"),
    "99VN": ("spare nose for item 99ST", "99ST için yedek uç"),
    "100": ("round pin wrench for ring nuts", "Segman somunları için yuvarlak pimli anahtar"),
    "100/KIT": ("spare pins for item 100", "100 için yedek pimler"),
    "111E": ("adjustable wrenches with scales, chrome-plated", "Ölçekli ayarlanabilir anahtarlar, krom kaplı"),
    "111EN": ("adjustable wrenches with scales, phosphated", "Ölçekli ayarlanabilir anahtarlar, fosfatlı"),
    "111ER": ("reversible jaw adjustable wrenches with scales, chrome-plated", "Çevrilebilir çeneli ölçekli ayarlanabilir anahtarlar, krom kaplı"),
    "111CM": ("wide opening adjustable wrenches, chrome-plated, short series", "Geniş açıklıklı ayarlanabilir anahtarlar, krom kaplı, kısa seri"),
    "120": ("ratchet opening single ended bi-hex wrenches", "Cırcırlı tek uçlu çift altıgen anahtarlar"),
    "123/K4": ("kit with 4 adapters for ratchet wrenches", "Cırcır anahtarları için 4 adaptörlü takım"),
    "123E1/4": ("bit holder adaptor, 1/4\", for 10 mm ratcheting wrenches", "Uç tutucu adaptör, 1/4\", 10 mm cırcırlı anahtarlar için"),
    "123Q1/4": ("quick release adaptor, 1/4\", for 10 mm ratcheting wrenches", "Hızlı çıkarma adaptörü, 1/4\", 10 mm cırcırlı anahtarlar için"),
    "123Q3/8": ("quick release adaptor, 3/8\", for 13 mm ratcheting wrenches", "Hızlı çıkarma adaptörü, 3/8\", 13 mm cırcırlı anahtarlar için"),
    "123Q1/2": ("quick release adaptor, 1/2\", for 19 mm ratcheting wrenches", "Hızlı çıkarma adaptörü, 1/2\", 19 mm cırcırlı anahtarlar için"),
    "141": ("ratcheting combination wrenches, straight series", "Cırcırlı kombinasyon anahtarları, düz seri"),
}

def get_product_name_tr(sku, product_code, size=""):
    """SKU ve ürün koduna göre Türkçe isim döndür"""
    # Önce tam eşleşme ara
    if product_code in product_names:
        _, tr_name = product_names[product_code]
        if size:
            return f"Beta {product_code} {tr_name} - {size}"
        return f"Beta {product_code} {tr_name}"
    
    # Kısmi eşleşme ara (örn: 42/S14 -> 42)
    for code in product_names:
        if product_code.startswith(code) or code in product_code:
            _, tr_name = product_names[code]
            if size:
                return f"Beta {product_code} {tr_name} - {size}"
            return f"Beta {product_code} {tr_name}"
    
    # Eşleşme bulunamadı
    if size:
        return f"Beta {product_code} - {size}"
    return f"Beta {product_code}"

def extract_all_data(pricelist_path, gp_path, max_pages=75):
    """Her iki PDF'den verileri çıkar ve birleştir"""
    products = []
    
    # Önce GP_ENG'den ürün kodları ve SKU eşleştirmesi çıkar
    sku_to_product = {}
    current_product_code = ""
    
    with pdfplumber.open(gp_path) as pdf:
        for page_num in range(9, min(max_pages, len(pdf.pages))):
            page = pdf.pages[page_num]
            text = page.extract_text()
            if not text:
                continue
            
            lines = text.split('\n')
            
            for line in lines:
                # Ürün kodu pattern'i (|* ile başlayan)
                code_match = re.search(r'\|?\*?\s*(\d+[A-Z\-/]*[A-Z0-9]*)', line)
                if code_match and '|*' in line:
                    potential_code = code_match.group(1).strip()
                    if potential_code and len(potential_code) >= 2:
                        current_product_code = potential_code
                
                # SKU pattern'i
                sku_matches = re.findall(r'(0\d{8})', line)
                for sku in sku_matches:
                    # Boyut bilgisi çıkar
                    # Format: boyut fiyat adet SKU
                    size_pattern = r'(\d+(?:[,x\.]\d+)*(?:x\d+(?:[,\.]\d+)*)?)\s+\d+[\.,]?\d*\s+\d+\s+' + sku
                    size_match = re.search(size_pattern, line)
                    size = size_match.group(1) if size_match else ""
                    
                    if sku not in sku_to_product:
                        sku_to_product[sku] = {
                            'product_code': current_product_code,
                            'size': size
                        }
    
    print(f"GP_ENG'den {len(sku_to_product)} SKU eşleştirmesi çıkarıldı")
    
    # PriceList'ten fiyatları çıkar
    with pdfplumber.open(pricelist_path) as pdf:
        for page_num in range(9, min(max_pages, len(pdf.pages))):
            page = pdf.pages[page_num]
            text = page.extract_text()
            if not text:
                continue
            
            # Fiyat ve SKU pattern'i
            pattern = r'(\d+\.?\d*)\s+(\d+)\s+(0\d{8})'
            matches = re.findall(pattern, text)
            
            for match in matches:
                price_gbp = float(match[0])
                quantity = int(match[1])
                sku = match[2]
                
                # Fiyatı 41 ile çarp
                price_try = round(price_gbp * 41, 2)
                
                # Ürün bilgisi al
                info = sku_to_product.get(sku, {})
                product_code = info.get('product_code', '')
                size = info.get('size', '')
                
                # Türkçe isim al
                product_name = get_product_name_tr(sku, product_code, size)
                
                # Açıklama oluştur
                desc_parts = [f"Beta {product_code}"]
                if size:
                    desc_parts.append(f"Boyut: {size}")
                desc_parts.append(f"GBP Fiyat: £{price_gbp}")
                description = "\n".join(desc_parts)
                
                products.append({
                    'StokKodu': sku,
                    'UrunAdi': product_name,
                    'Marka': 'Beta',
                    'Fiyat': price_try,
                    'IndirimliFiyat': '',
                    'Stok': 50,
                    'Kategori': 'Hırdavat ve El Aletleri',
                    'AltKategori': 'El Aletleri',
                    'Aciklama': description,
                    'Birim': 'adet',
                    'GorselURL': f"Beta_Katalog_Gorseller_Final/{sku}.jpg",
                    'Aktif': 'Evet',
                    'PopulerMi': 'Hayır',
                    'YeniMi': 'Evet',
                    'OneCikan': 'Hayır',
                    'CokSatan': 'Hayır',
                    'MarkaVitrini': ''
                })
    
    return products

def create_excel(products, output_path):
    """Excel dosyası oluştur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ürünler"
    
    # Başlıklar (Lidareyn formatı)
    headers = ['StokKodu', 'UrunAdi', 'Marka', 'Fiyat', 'IndirimliFiyat', 'Stok', 
               'Kategori', 'AltKategori', 'Aciklama', 'Birim', 'GorselURL', 
               'Aktif', 'PopulerMi', 'YeniMi', 'OneCikan', 'CokSatan', 'MarkaVitrini']
    
    # Başlık stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlıkları yaz
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Verileri yaz
    for row_num, product in enumerate(products, 2):
        for col, header in enumerate(headers, 1):
            value = product.get(header, '')
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
    
    # Sütun genişliklerini ayarla
    column_widths = {
        'A': 15,  # StokKodu
        'B': 70,  # UrunAdi
        'C': 10,  # Marka
        'D': 12,  # Fiyat
        'E': 15,  # IndirimliFiyat
        'F': 8,   # Stok
        'G': 25,  # Kategori
        'H': 20,  # AltKategori
        'I': 40,  # Aciklama
        'J': 8,   # Birim
        'K': 45,  # GorselURL
        'L': 8,   # Aktif
        'M': 10,  # PopulerMi
        'N': 8,   # YeniMi
        'O': 10,  # OneCikan
        'P': 10,  # CokSatan
        'Q': 12,  # MarkaVitrini
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Satır 1'i dondur
    ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    print(f"Excel dosyası kaydedildi: {output_path}")
    return len(products)

# Ana çalıştırma
if __name__ == "__main__":
    pricelist_path = "/mnt/uploads/ses_3b3d28946ffe2yh5rSr2HGXvow/PriceList_2025_GBP.pdf"
    gp_path = "/mnt/workspace/258iT5oxNEFM9USsyJp4GKR9LsRrdMyu63pcuFxepD6i7PR/GP_ENG_2025.pdf"
    output_path = "/mnt/workspace/258iT5oxNEFM9USsyJp4GKR9LsRrdMyu63pcuFxepD6i7PR/Beta_Urunler_2026.xlsx"
    
    print("Veriler çıkarılıyor...")
    products = extract_all_data(pricelist_path, gp_path, 75)
    print(f"Toplam {len(products)} ürün bulundu")
    
    print("\nExcel oluşturuluyor...")
    count = create_excel(products, output_path)
    print(f"\nTamamlandı! {count} ürün Excel'e yazıldı.")
    
    # İlk 10 ürünü göster
    print("\nÖrnek ürünler:")
    for p in products[:10]:
        print(f"SKU: {p['StokKodu']}, Fiyat: {p['Fiyat']} TRY")
        print(f"  Ürün: {p['UrunAdi']}")
        print()
