#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GP_ENG_2025.pdf'den ürün bilgilerini ve PriceList_2025_GBP.pdf'den fiyatları çıkarır
Excel formatına dönüştürür - Final versiyon
"""

import pdfplumber
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Ürün kodu -> Türkçe isim eşleştirmesi
product_names_tr = {
    "32AUR": "Otomatik zımba yedek uç",
    "33LT": "Harf zımba seti",
    "33NR": "Rakam zımba seti",
    "34": "Düz keskiler",
    "35": "Düz keskiler, nervürlü tip",
    "35PM": "Düz keskiler, nervürlü tip, el korumalı",
    "35PMR": "Yedek el korumaları",
    "36": "Sivri uçlu keskiler",
    "37": "Sivri uçlu keskiler, nervürlü tip",
    "37PM": "Sivri uçlu keski, nervürlü tip, el korumalı",
    "37PMR": "Yedek el koruması",
    "38/B6N": "Keski seti",
    "38/SP6": "Yedek keski uçları",
    "38/SPV": "Yedek keski uçları",
    "42": "Kombinasyon anahtarı",
    "42AS": "Kombinasyon anahtarı, açık ve ofset halka uçlu",
    "42MP": "Kombinasyon anahtarı, parlak krom kaplı",
    "42LMP": "Kombinasyon anahtarı, uzun seri, parlak krom kaplı",
    "42SLIM": "İnce açık uçlu kombinasyon anahtarı",
    "42/SC9I": "9'lu kombinasyon anahtar seti",
    "42/SC9E": "9'lu kombinasyon anahtar seti",
    "42/SP15": "15'li kombinasyon anahtar seti",
    "42/SP25": "25'li kombinasyon anahtar seti",
    "42/S14": "14'lü kombinasyon anahtar seti",
    "42/S15": "15'li kombinasyon anahtar seti",
    "42/S17": "17'li kombinasyon anahtar seti",
    "42/S17X": "17'li kombinasyon anahtar seti",
    "42/S26": "26'lı kombinasyon anahtar seti",
    "42/SPV7": "7'li kombinasyon anahtar seti",
    "42/SPV11": "11'li kombinasyon anahtar seti",
    "42/SPV15": "15'li kombinasyon anahtar seti",
    "42/SPV2": "2'li kombinasyon anahtar seti",
    "42AS/B9N": "9'lu kombinasyon anahtar seti",
    "42AS/13": "13'lü kombinasyon anahtar seti",
    "42AS/16": "16'lı kombinasyon anahtar seti",
    "42AS/17": "17'li kombinasyon anahtar seti",
    "42MP/B9": "9'lu kombinasyon anahtar seti",
    "42MP/S17": "17'li kombinasyon anahtar seti",
    "42MP/S21": "21'li kombinasyon anahtar seti",
    "42LMP/S14": "14'lü kombinasyon anahtar seti",
    "42LMP/S18": "18'li kombinasyon anahtar seti",
    "42SLIM/S11": "11'li ince kombinasyon anahtar seti",
    "42SLIM/B8NI": "8'li ince kombinasyon anahtar seti",
    "45": "Kombinasyon anahtarı, ağır seri",
    "52": "Tek açık uçlu anahtar",
    "53": "Tek açık uçlu anahtar",
    "55": "Çift açık uçlu anahtar",
    "55AS": "Çift açık uçlu anahtar",
    "55/B8N": "8'li çift açık uçlu anahtar seti",
    "55/B12N": "12'li çift açık uçlu anahtar seti",
    "55/S8": "8'li çift açık uçlu anahtar seti",
    "55/S12": "12'li çift açık uçlu anahtar seti",
    "55/S12X": "12'li çift açık uçlu anahtar seti",
    "55/S13": "13'lü çift açık uçlu anahtar seti",
    "55/SP7": "7'li çift açık uçlu anahtar seti",
    "55/SP13": "13'lü çift açık uçlu anahtar seti",
    "55/SPV": "Çift açık uçlu anahtar seti",
    "55AS/5": "5'li çift açık uçlu anahtar seti",
    "55AS/8": "8'li çift açık uçlu anahtar seti",
    "55AS/12": "12'li çift açık uçlu anahtar seti",
    "58": "Açık uçlu balyoz anahtarı",
    "73": "Küçük çift açık uçlu anahtar",
    "73/B13N": "13'lü küçük çift açık uçlu anahtar seti",
    "78": "Halka balyoz anahtarı",
    "78AS": "Halka balyoz anahtarı",
    "80": "Çift mafsallı lokma anahtarı",
    "80/S8": "8'li çift mafsallı lokma anahtar seti",
    "80/S11": "11'li çift mafsallı lokma anahtar seti",
    "83": "Yarım ay halka anahtarı",
    "83AS": "Yarım ay halka anahtarı",
    "83AS/3": "3'lü yarım ay halka anahtar seti",
    "88": "Çift uçlu düz halka anahtarı, ekstra uzun seri",
    "88/S7": "7'li çift uçlu düz halka anahtar seti",
    "90": "Çift uçlu derin ofset halka anahtarı",
    "90AS": "Çift uçlu derin ofset halka anahtarı",
    "90/B8N": "8'li çift uçlu derin ofset halka anahtar seti",
    "90/S8": "8'li çift uçlu derin ofset halka anahtar seti",
    "90/S12": "12'li çift uçlu derin ofset halka anahtar seti",
    "90/S12X": "12'li çift uçlu derin ofset halka anahtar seti",
    "90/S13": "13'lü çift uçlu derin ofset halka anahtar seti",
    "90/SP13": "13'lü çift uçlu derin ofset halka anahtar seti",
    "90AS/8": "8'li çift uçlu derin ofset halka anahtar seti",
    "90AS/10": "10'lu çift uçlu derin ofset halka anahtar seti",
    "91": "Ağır hizmet ofset halka anahtarı",
    "92": "91 için boru",
    "92/1": "91 için boru",
    "92/2": "91 için boru",
    "92/3": "91 için boru",
    "92/4": "91 için boru",
    "93": "İskele cıvataları için çift uçlu ofset halka anahtar",
    "93C": "İskele cıvataları için cırcırlı çift uçlu ofset halka anahtar",
    "94": "Fren borusu anahtarı",
    "94/S6": "6'lı fren borusu anahtar seti",
    "94/S10": "10'lu fren borusu anahtar seti",
    "95": "Çift uçlu düz halka anahtarı",
    "95/S8": "8'li çift uçlu düz halka anahtar seti",
    "95/S13": "13'lü çift uçlu düz halka anahtar seti",
    "95FTX": "Torx başlı vidalar için çift uçlu düz anahtar",
    "95FTX/S4": "4'lü Torx çift uçlu düz anahtar seti",
    "96": "Ofset altıgen anahtar, krom kaplı",
    "96N": "Ofset altıgen anahtar, parlatılmış",
    "96AS": "Ofset altıgen anahtar, parlatılmış",
    "96LC": "Ofset altıgen anahtar, uzun seri, krom kaplı",
    "96L": "Ofset altıgen anahtar, uzun seri, parlatılmış",
    "96BP": "Bilyalı uç ofset altıgen anahtar, parlatılmış",
    "96BP-CL": "Bilyalı uç ofset altıgen anahtar, krom kaplı, renkli",
    "96BPC": "Bilyalı uç ofset altıgen anahtar, krom kaplı",
    "96LBP": "Bilyalı uç ofset altıgen anahtar, ekstra uzun model",
    "96BP-HO": "Vida tutma sistemli bilyalı uç ofset altıgen anahtar",
    "96T": "Yüksek torklu saplarla ofset altıgen anahtar",
    "96T/AS": "Yüksek torklu saplarla ofset altıgen anahtar",
    "96TBP": "Yüksek torklu saplarla bilyalı uç ofset altıgen anahtar",
    "96BPA": "Bilyalı uç ofset altıgen anahtar, 110°, ekstra kısa yan model",
    "96/B8": "8'li ofset altıgen anahtar seti",
    "96/B10": "10'lu ofset altıgen anahtar seti",
    "96/SC9": "9'lu ofset altıgen anahtar seti",
    "96/SC12": "12'li ofset altıgen anahtar seti",
    "96/ST6": "6'lı ofset altıgen anahtar seti",
    "96/BGS7": "7'li ofset altıgen anahtar seti",
    "96/BG7": "7'li ofset altıgen anahtar seti",
    "96N/ST6": "6'lı ofset altıgen anahtar seti",
    "96N/G7": "7'li ofset altıgen anahtar seti",
    "96N/G7A": "7'li ofset altıgen anahtar seti",
    "96N/B10": "10'lu ofset altıgen anahtar seti",
    "96N/SC9": "9'lu ofset altıgen anahtar seti",
    "96N/SC12": "12'li ofset altıgen anahtar seti",
    "96N/BV": "Ofset altıgen anahtar tutucu",
    "96/SCV": "Ofset altıgen anahtar tutucu",
    "96/SCV12": "12'li ofset altıgen anahtar tutucu",
    "96N/SCV": "Ofset altıgen anahtar tutucu",
    "96N/SCV12": "12'li ofset altıgen anahtar tutucu",
    "96AS/SC9": "9'lu ofset altıgen anahtar seti",
    "96AS/SCV": "Ofset altıgen anahtar tutucu",
    "96AS/T7": "7'li ofset altıgen anahtar seti",
    "96AS/BG7": "7'li ofset altıgen anahtar seti",
    "96AS/B11": "11'li ofset altıgen anahtar seti",
    "96LC/SC8": "8'li ofset altıgen anahtar seti",
    "96LC/SC12": "12'li ofset altıgen anahtar seti",
    "96LC/SCV": "Ofset altıgen anahtar tutucu",
    "96LC/SCV12": "12'li ofset altıgen anahtar tutucu",
    "96L/SC8": "8'li ofset altıgen anahtar seti",
    "96L/SC12": "12'li ofset altıgen anahtar seti",
    "96L/SCV": "Ofset altıgen anahtar tutucu",
    "96L/SCV12": "12'li ofset altıgen anahtar tutucu",
    "96BP/SC9": "9'lu bilyalı uç ofset altıgen anahtar seti",
    "96BP/SCV": "Bilyalı uç ofset altıgen anahtar tutucu",
    "96BP-CL/SC9": "9'lu bilyalı uç ofset altıgen anahtar seti, renkli",
    "96BP-CL/SCV": "Bilyalı uç ofset altıgen anahtar tutucu, renkli",
    "96BP-HO/SC9": "9'lu vida tutma sistemli bilyalı uç ofset altıgen anahtar seti",
    "96BPC/SC9": "9'lu bilyalı uç ofset altıgen anahtar seti",
    "96BPC/SCV": "Bilyalı uç ofset altıgen anahtar tutucu",
    "96LBP/SC9": "9'lu bilyalı uç ofset altıgen anahtar seti, ekstra uzun",
    "96LBP/SCV": "Bilyalı uç ofset altıgen anahtar tutucu, ekstra uzun",
    "96BPA/SC9": "9'lu bilyalı uç ofset altıgen anahtar seti, 110°",
    "96BPA/SCV": "Bilyalı uç ofset altıgen anahtar tutucu, 110°",
    "96T/S5P": "5'li yüksek torklu ofset altıgen anahtar seti",
    "96T/S6": "6'lı yüksek torklu ofset altıgen anahtar seti",
    "96T/S8": "8'li yüksek torklu ofset altıgen anahtar seti",
    "96T/S11": "11'li yüksek torklu ofset altıgen anahtar seti",
    "96T/SP11": "11'li yüksek torklu ofset altıgen anahtar seti",
    "96T/SPV": "Yüksek torklu ofset altıgen anahtar tutucu",
    "96T/AS8": "8'li yüksek torklu ofset altıgen anahtar seti",
    "96T/AS10": "10'lu yüksek torklu ofset altıgen anahtar seti",
    "96TBP/S6": "6'lı yüksek torklu bilyalı uç ofset altıgen anahtar seti",
    "97TX": "Torx başlı vidalar için ofset anahtar",
    "97BTX": "Torx başlı vidalar için bilyalı uç ofset anahtar",
    "97BTXL": "Torx başlı vidalar için bilyalı uç ofset anahtar, uzun model",
    "97RTX": "Güvenlikli Torx başlı vidalar için ofset anahtar",
    "97BRTXL": "Güvenlikli Torx başlı vidalar için bilyalı uç ofset anahtar, uzun model",
    "97TTX": "Torx başlı vidalar için saplı ofset anahtar",
    "97TX/ST8": "8'li Torx ofset anahtar seti",
    "97TX/SC8": "8'li Torx ofset anahtar seti",
    "97TX/SCV": "Torx ofset anahtar tutucu",
    "97TX/G8": "8'li Torx ofset anahtar seti",
    "97TX/BG8": "8'li Torx ofset anahtar seti",
    "97TX/B13": "13'lü Torx ofset anahtar seti",
    "97BTX/SC8": "8'li bilyalı uç Torx ofset anahtar seti",
    "97BTX/SCV": "Bilyalı uç Torx ofset anahtar tutucu",
    "97BTX/B8": "8'li bilyalı uç Torx ofset anahtar seti",
    "97BTX-C/SC8": "8'li bilyalı uç Torx ofset anahtar seti, renkli",
    "97BTX-C/SCV": "Bilyalı uç Torx ofset anahtar tutucu, renkli",
    "97BTXL/SC8": "8'li bilyalı uç Torx ofset anahtar seti, uzun",
    "97BTXL/SCV": "Bilyalı uç Torx ofset anahtar tutucu, uzun",
    "97RTX/SC8": "8'li güvenlikli Torx ofset anahtar seti",
    "97RTX/SCV": "Güvenlikli Torx ofset anahtar tutucu",
    "97RTX/B8": "8'li güvenlikli Torx ofset anahtar seti",
    "97BRTXL/SC8": "8'li güvenlikli bilyalı uç Torx ofset anahtar seti, uzun",
    "97BRTXL/SCV": "Güvenlikli bilyalı uç Torx ofset anahtar tutucu, uzun",
    "97TTX/S6": "6'lı saplı Torx ofset anahtar seti",
    "97TTX/S8": "8'li saplı Torx ofset anahtar seti",
    "97TTX/S11": "11'li saplı Torx ofset anahtar seti",
    "97TTX/S13": "13'lü saplı Torx ofset anahtar seti",
    "97TTX/SP11": "11'li saplı Torx ofset anahtar seti",
    "98XZN": "XZN profilli ofset anahtar",
    "98XZN/B5": "5'li XZN profilli ofset anahtar seti",
    "99": "Segman somunları için kare uçlu kanca anahtar",
    "99SQ": "Segman somunları için kare uçlu kanca anahtar",
    "99ST": "Segman somunları için yuvarlak uçlu kanca anahtar",
    "99VN": "99ST için yedek uç",
    "99ST VN/15": "99ST için yedek uç 15-35mm",
    "99ST VN/35": "99ST için yedek uç 35-50mm",
    "99ST VN/50": "99ST için yedek uç 50-80mm",
    "99ST VN/80": "99ST için yedek uç 80-120mm",
    "99ST VN/120": "99ST için yedek uç 120-180mm",
    "100": "Segman somunları için yuvarlak pimli anahtar",
    "100/KIT": "100 için yedek pim seti",
    "111E": "Ölçekli ayarlanabilir anahtar, krom kaplı",
    "111EN": "Ölçekli ayarlanabilir anahtar, fosfatlı",
    "111ER": "Çevrilebilir çeneli ölçekli ayarlanabilir anahtar, krom kaplı",
    "111CM": "Geniş açıklıklı ayarlanabilir anahtar, krom kaplı, kısa seri",
    "111E/D3": "3'lü ölçekli ayarlanabilir anahtar seti",
    "120": "Cırcırlı tek uçlu çift altıgen anahtar",
    "120/B6N": "6'lı cırcırlı tek uçlu çift altıgen anahtar seti",
    "123/K4": "Cırcır anahtarları için 4 adaptörlü takım",
    "123E1/4": "Uç tutucu adaptör, 1/4\"",
    "123Q1/4": "Hızlı çıkarma adaptörü, 1/4\"",
    "123Q3/8": "Hızlı çıkarma adaptörü, 3/8\"",
    "123Q1/2": "Hızlı çıkarma adaptörü, 1/2\"",
    "141": "Cırcırlı kombinasyon anahtarı, düz seri",
    "141/B6": "6'lı cırcırlı kombinasyon anahtar seti",
    "141/B9": "9'lu cırcırlı kombinasyon anahtar seti",
    "141/S12": "12'li cırcırlı kombinasyon anahtar seti",
    "142": "Cırcırlı kombinasyon anahtarı",
    "142/SCV9I": "9'lu cırcırlı kombinasyon anahtar seti",
    "142/SCV9E": "9'lu cırcırlı kombinasyon anahtar seti",
    "2002/BV6": "6'lı anahtar tutucu",
    "2002/BV8": "8'li anahtar tutucu",
    "2002/BV9": "9'lu anahtar tutucu",
    "2002/BV12": "12'li anahtar tutucu",
    "2002/BV13": "13'lü anahtar tutucu",
    "2002/BVL6": "6'lı uzun anahtar tutucu",
}

def get_product_name(product_code, size=""):
    """Ürün koduna göre Türkçe isim döndür"""
    # Tam eşleşme ara
    if product_code in product_names_tr:
        name = product_names_tr[product_code]
        if size:
            return f"Beta {product_code} {name} - {size}"
        return f"Beta {product_code} {name}"
    
    # Kısmi eşleşme ara
    for code in sorted(product_names_tr.keys(), key=len, reverse=True):
        if product_code.startswith(code) or code in product_code:
            name = product_names_tr[code]
            if size:
                return f"Beta {product_code} {name} - {size}"
            return f"Beta {product_code} {name}"
    
    # Eşleşme bulunamadı
    if size:
        return f"Beta {product_code} - {size}"
    return f"Beta {product_code}"

def extract_all_data(pricelist_path, gp_path, max_pages=75):
    """Her iki PDF'den verileri çıkar ve birleştir"""
    products = []
    
    # GP_ENG'den ürün kodları ve SKU eşleştirmesi çıkar
    sku_to_product = {}
    current_product_code = ""
    
    with pdfplumber.open(gp_path) as pdf:
        for page_num in range(9, min(max_pages, len(pdf.pages))):
            page = pdf.pages[page_num]
            text = page.extract_text()
            if not text:
                continue
            
            lines = text.split('\n')
            
            for i, line in enumerate(lines):
                # |* işaretinden sonraki satır ürün kodu
                if '|*' in line or line.strip() == '||**':
                    # Sonraki satırda ürün kodu olabilir
                    if i + 1 < len(lines):
                        next_line = lines[i + 1].strip()
                        # Ürün kodu pattern'i (rakam veya rakam+harf kombinasyonu)
                        if re.match(r'^[\d]+[A-Z/\-]*[A-Z0-9]*$', next_line) and len(next_line) >= 2:
                            current_product_code = next_line
                
                # SKU pattern'i
                sku_matches = re.findall(r'(0\d{8})', line)
                for sku in sku_matches:
                    # Boyut bilgisi çıkar (SKU'dan önceki sayısal değer)
                    # Format: boyut ... SKU veya boyutxboyut ... SKU
                    size = ""
                    size_patterns = [
                        r'(\d+(?:[,\.]\d+)?x\d+(?:[,\.]\d+)?)\s+.*?' + sku,  # 10x11, 6,5x7
                        r'(\d+/\d+x\d+/\d+)\s+.*?' + sku,  # 1/4x1/4
                        r'(\d+(?:[,\.]\d+)?)\s+\d+[\.,]?\d*\s+\d+\s+' + sku,  # 100 14,5 5 SKU
                    ]
                    
                    for pattern in size_patterns:
                        match = re.search(pattern, line)
                        if match:
                            size = match.group(1)
                            break
                    
                    if sku not in sku_to_product:
                        sku_to_product[sku] = {
                            'product_code': current_product_code,
                            'size': size
                        }
    
    print(f"GP_ENG'den {len(sku_to_product)} SKU eşleştirmesi çıkarıldı")
    
    # PriceList'ten fiyatları çıkar
    with pdfplumber.open(pricelist_path) as pdf:
        for page_num in range(9, min(max_pages, len(pdf.pages))):
            page = pdf.pages[page_num]
            text = page.extract_text()
            if not text:
                continue
            
            # Fiyat ve SKU pattern'i
            pattern = r'(\d+\.?\d*)\s+(\d+)\s+(0\d{8})'
            matches = re.findall(pattern, text)
            
            for match in matches:
                price_gbp = float(match[0])
                quantity = int(match[1])
                sku = match[2]
                
                # Fiyatı 41 ile çarp
                price_try = round(price_gbp * 41, 2)
                
                # Ürün bilgisi al
                info = sku_to_product.get(sku, {})
                product_code = info.get('product_code', '')
                size = info.get('size', '')
                
                # Türkçe isim al
                product_name = get_product_name(product_code, size)
                
                # Açıklama oluştur
                desc_parts = [f"Beta {product_code}"]
                if size:
                    desc_parts.append(f"Boyut: {size}")
                desc_parts.append(f"GBP Fiyat: £{price_gbp}")
                description = "\n".join(desc_parts)
                
                products.append({
                    'StokKodu': sku,
                    'UrunAdi': product_name,
                    'Marka': 'Beta',
                    'Fiyat': price_try,
                    'IndirimliFiyat': '',
                    'Stok': 50,
                    'Kategori': 'Hırdavat ve El Aletleri',
                    'AltKategori': 'El Aletleri',
                    'Aciklama': description,
                    'Birim': 'adet',
                    'GorselURL': f"Beta_Katalog_Gorseller_Final/{sku}.jpg",
                    'Aktif': 'Evet',
                    'PopulerMi': 'Hayır',
                    'YeniMi': 'Evet',
                    'OneCikan': 'Hayır',
                    'CokSatan': 'Hayır',
                    'MarkaVitrini': ''
                })
    
    return products

def create_excel(products, output_path):
    """Excel dosyası oluştur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ürünler"
    
    # Başlıklar
    headers = ['StokKodu', 'UrunAdi', 'Marka', 'Fiyat', 'IndirimliFiyat', 'Stok', 
               'Kategori', 'AltKategori', 'Aciklama', 'Birim', 'GorselURL', 
               'Aktif', 'PopulerMi', 'YeniMi', 'OneCikan', 'CokSatan', 'MarkaVitrini']
    
    # Başlık stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlıkları yaz
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Verileri yaz
    for row_num, product in enumerate(products, 2):
        for col, header in enumerate(headers, 1):
            value = product.get(header, '')
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
    
    # Sütun genişlikleri
    column_widths = {'A': 15, 'B': 70, 'C': 10, 'D': 12, 'E': 15, 'F': 8, 
                     'G': 25, 'H': 20, 'I': 40, 'J': 8, 'K': 45, 'L': 8, 
                     'M': 10, 'N': 8, 'O': 10, 'P': 10, 'Q': 12}
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = 'A2'
    wb.save(output_path)
    return len(products)

# Ana çalıştırma
if __name__ == "__main__":
    pricelist_path = "/mnt/uploads/ses_3b3d28946ffe2yh5rSr2HGXvow/PriceList_2025_GBP.pdf"
    gp_path = "/mnt/workspace/258iT5oxNEFM9USsyJp4GKR9LsRrdMyu63pcuFxepD6i7PR/GP_ENG_2025.pdf"
    output_path = "/mnt/workspace/258iT5oxNEFM9USsyJp4GKR9LsRrdMyu63pcuFxepD6i7PR/Beta_Urunler_2026.xlsx"
    
    print("Veriler çıkarılıyor...")
    products = extract_all_data(pricelist_path, gp_path, 75)
    print(f"Toplam {len(products)} ürün bulundu")
    
    print("\nExcel oluşturuluyor...")
    count = create_excel(products, output_path)
    print(f"\nTamamlandı! {count} ürün Excel'e yazıldı.")
    print(f"Dosya: {output_path}")
    
    # İlk 15 ürünü göster
    print("\nÖrnek ürünler:")
    for p in products[:15]:
        print(f"SKU: {p['StokKodu']}, Fiyat: {p['Fiyat']} TRY")
        print(f"  {p['UrunAdi']}")
