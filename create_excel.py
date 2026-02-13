#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GP_ENG_2025.pdf'den ürün bilgilerini ve PriceList_2025_GBP.pdf'den fiyatları çıkarır
Excel formatına dönüştürür
"""

import pdfplumber
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# İngilizce-Türkçe çeviri sözlüğü
translations = {
    "spare tip for automatic centre punch": "Otomatik zımba yedek uç",
    "flat chisels, ribbed type": "Düz keskiler, nervürlü tip",
    "cape chisel, ribbed type": "Sivri uçlu keskiler, nervürlü tip",
    "with hand guard": "el korumalı",
    "letter punches": "Harf zımbaları",
    "number punches": "Rakam zımbaları",
    "spare hand guard": "Yedek el koruması",
    "spare hand guards": "Yedek el korumaları",
    "cape chisels": "Sivri uçlu keskiler",
    "combination wrenches": "Kombinasyon anahtarları",
    "open and offset ring ends": "Açık ve ofset halka uçlu",
    "bright chrome-plated": "Parlak krom kaplı",
    "long series": "Uzun seri",
    "heavy series": "Ağır seri",
    "single open end wrenches": "Tek açık uçlu anahtarlar",
    "double open end wrenches": "Çift açık uçlu anahtarlar",
    "open end slogging wrenches": "Açık uçlu balyoz anahtarları",
    "ring slogging wrenches": "Halka balyoz anahtarları",
    "double swivel end socket wrenches": "Çift mafsallı lokma anahtarları",
    "half-moon ring wrenches": "Yarım ay halka anahtarları",
    "double ended deep offset ring wrenches": "Çift uçlu derin ofset halka anahtarları",
    "double ended flat ring wrenches": "Çift uçlu düz halka anahtarları",
    "extra-long series": "Ekstra uzun seri",
    "heavy duty offset ring wrenches": "Ağır hizmet ofset halka anahtarları",
    "flare nut open ring wrenches": "Fren borusu anahtarları",
    "double-ended straight wrenches": "Çift uçlu düz anahtarlar",
    "for Torx head screws": "Torx başlı vidalar için",
    "for Torx® head screws": "Torx başlı vidalar için",
    "offset hexagon key wrenches": "Ofset altıgen anahtar",
    "chrome-plated": "Krom kaplı",
    "burnished": "Parlatılmış",
    "ball head": "Bilyalı uç",
    "with high torque handles": "Yüksek torklu saplar ile",
    "hook wrenches with square noses": "Kare uçlu kanca anahtarlar",
    "hook wrenches with round noses": "Yuvarlak uçlu kanca anahtarlar",
    "for ring nuts": "Segman somunları için",
    "round pin wrench": "Yuvarlak pimli anahtar",
    "spare pins": "Yedek pimler",
    "adjustable wrenches with scales": "Ölçekli ayarlanabilir anahtarlar",
    "wide opening": "Geniş açıklık",
    "short series": "Kısa seri",
    "reversible jaw": "Çevrilebilir çene",
    "ratchet opening single ended": "Cırcırlı tek uçlu",
    "bi-hex wrenches": "Çift altıgen anahtarlar",
    "bit holder adaptor": "Uç tutucu adaptör",
    "quick release adaptor": "Hızlı çıkarma adaptörü",
    "ratcheting combination wrenches": "Cırcırlı kombinasyon anahtarları",
    "straight series": "Düz seri",
    "small double open end wrenches": "Küçük çift açık uçlu anahtarlar",
    "tubes": "Borular",
    "double-ended offset ring wrench": "Çift uçlu ofset halka anahtar",
    "for scaffolding bolts": "İskele cıvataları için",
    "ratcheting double-ended offset ring wrenches": "Cırcırlı çift uçlu ofset halka anahtarlar",
    "with screw holding system": "Vida tutma sistemi ile",
    "extra-long model": "Ekstra uzun model",
    "coloured": "Renkli",
    "for Tamper Resistant Torx head screws": "Güvenlikli Torx başlı vidalar için",
    "for Tamper Resistant Torx® head screws": "Güvenlikli Torx başlı vidalar için",
    "offset key wrenches with handles": "Saplı ofset anahtarlar",
    "offset key wrenches": "Ofset anahtarlar",
    "with XZN profile": "XZN profilli",
    "with XZN® profile": "XZN profilli",
    "spare nose": "Yedek uç",
    "adapters for ratchet wrenches": "Cırcır anahtarları için adaptörler",
    "kit with": "Takım",
    "set of": "Takım",
    "wrenches": "anahtarlar",
    "wrench": "anahtar",
    "screwdrivers": "tornavidalar",
    "screwdriver": "tornavida",
    "pliers": "pense",
    "hammers": "çekiçler",
    "hammer": "çekiç",
    "chisels": "keskiler",
    "chisel": "keski",
    "bits": "uçlar",
    "bit": "uç",
    "sockets": "lokmalar",
    "socket": "lokma",
    "extensions": "uzatmalar",
    "extension": "uzatma",
    "ratchets": "cırcırlar",
    "ratchet": "cırcır",
    "torque": "tork",
    "impact": "darbe",
    "insulated": "yalıtımlı",
    "magnetic": "manyetik",
    "universal": "üniversal",
    "professional": "profesyonel",
    "precision": "hassas",
    "standard": "standart",
    "compact": "kompakt",
    "support": "destek",
    "phosphated": "fosfatlı",
    "reversible": "çevrilebilir",
    "jaw": "çene",
    "wide": "geniş",
    "opening": "açıklık",
}

def translate_name(english_name):
    """İngilizce ürün ismini Türkçeye çevir"""
    result = english_name
    # Önce uzun ifadeleri çevir
    sorted_translations = sorted(translations.items(), key=lambda x: len(x[0]), reverse=True)
    for eng, tr in sorted_translations:
        result = re.sub(re.escape(eng), tr, result, flags=re.IGNORECASE)
    return result

def extract_all_data(pricelist_path, gp_path, max_pages=75):
    """Her iki PDF'den verileri çıkar ve birleştir"""
    products = []
    
    # Önce GP_ENG'den ürün isimlerini ve yapısını çıkar
    product_info = {}
    current_product_name = ""
    current_product_code = ""
    
    with pdfplumber.open(gp_path) as pdf:
        for page_num in range(9, min(max_pages, len(pdf.pages))):
            page = pdf.pages[page_num]
            text = page.extract_text()
            if not text:
                continue
            
            lines = text.split('\n')
            
            for line in lines:
                # Ürün başlığı pattern'i (|* ile başlayan satırlar genellikle ürün kodu)
                if '|*' in line or line.strip().startswith('*'):
                    code_match = re.search(r'\|?\*?\s*(\d+[A-Z]*/?[A-Z0-9\-]*)', line)
                    if code_match:
                        current_product_code = code_match.group(1).strip()
                
                # Ürün açıklaması (küçük harfle başlayan satırlar)
                if re.match(r'^[a-z]', line.strip()) and len(line.strip()) > 10:
                    current_product_name = line.strip()
                
                # SKU pattern'i (9 haneli sayı)
                sku_matches = re.findall(r'(0\d{8})', line)
                for sku in sku_matches:
                    if sku not in product_info:
                        # Boyut bilgisi çıkar
                        size_match = re.search(r'(\d+(?:[,x\.]\d+)*(?:x\d+(?:[,\.]\d+)*)?)\s+', line)
                        size = size_match.group(1) if size_match else ""
                        
                        product_info[sku] = {
                            'product_code': current_product_code,
                            'product_name': current_product_name,
                            'size': size
                        }
    
    print(f"GP_ENG'den {len(product_info)} ürün bilgisi çıkarıldı")
    
    # PriceList'ten fiyatları çıkar
    with pdfplumber.open(pricelist_path) as pdf:
        for page_num in range(9, min(max_pages, len(pdf.pages))):
            page = pdf.pages[page_num]
            text = page.extract_text()
            if not text:
                continue
            
            # Fiyat ve SKU pattern'i
            # Format: fiyat adet SKU
            pattern = r'(\d+\.?\d*)\s+(\d+)\s+(0\d{8})'
            matches = re.findall(pattern, text)
            
            for match in matches:
                price_gbp = float(match[0])
                quantity = int(match[1])
                sku = match[2]
                
                # Fiyatı 41 ile çarp
                price_try = round(price_gbp * 41, 2)
                
                # GP_ENG'den ürün bilgisi al
                info = product_info.get(sku, {})
                product_code = info.get('product_code', '')
                product_name_en = info.get('product_name', '')
                size = info.get('size', '')
                
                # Türkçeye çevir
                product_name_tr = translate_name(product_name_en) if product_name_en else ""
                
                # Ürün adı oluştur
                if product_code and size:
                    full_name = f"Beta {product_code} {product_name_tr} - {size}"
                elif product_code:
                    full_name = f"Beta {product_code} {product_name_tr}"
                else:
                    full_name = f"Beta {product_name_tr}" if product_name_tr else f"Beta Ürün {sku}"
                
                products.append({
                    'StokKodu': sku,
                    'UrunAdi': full_name.strip(),
                    'Marka': 'Beta',
                    'Fiyat': price_try,
                    'IndirimliFiyat': '',
                    'Stok': 50,
                    'Kategori': 'Hırdavat ve El Aletleri',
                    'AltKategori': '',
                    'Aciklama': f"Beta {product_code}\nBoyut: {size}\nGBP Fiyat: £{price_gbp}" if size else f"Beta {product_code}\nGBP Fiyat: £{price_gbp}",
                    'Birim': 'adet',
                    'GorselURL': f"Beta_Katalog_Gorseller_Final/{sku}.jpg",
                    'Aktif': 'Evet',
                    'PopulerMi': 'Hayır',
                    'YeniMi': 'Evet',
                    'OneCikan': 'Hayır',
                    'CokSatan': 'Hayır',
                    'MarkaVitrini': ''
                })
    
    return products

def create_excel(products, output_path):
    """Excel dosyası oluştur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ürünler"
    
    # Başlıklar (Lidareyn formatı)
    headers = ['StokKodu', 'UrunAdi', 'Marka', 'Fiyat', 'IndirimliFiyat', 'Stok', 
               'Kategori', 'AltKategori', 'Aciklama', 'Birim', 'GorselURL', 
               'Aktif', 'PopulerMi', 'YeniMi', 'OneCikan', 'CokSatan', 'MarkaVitrini']
    
    # Başlık stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Başlıkları yaz
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Verileri yaz
    for row_num, product in enumerate(products, 2):
        for col, header in enumerate(headers, 1):
            value = product.get(header, '')
            ws.cell(row=row_num, column=col, value=value)
    
    # Sütun genişliklerini ayarla
    column_widths = {
        'A': 15,  # StokKodu
        'B': 60,  # UrunAdi
        'C': 10,  # Marka
        'D': 12,  # Fiyat
        'E': 15,  # IndirimliFiyat
        'F': 8,   # Stok
        'G': 25,  # Kategori
        'H': 25,  # AltKategori
        'I': 50,  # Aciklama
        'J': 8,   # Birim
        'K': 40,  # GorselURL
        'L': 8,   # Aktif
        'M': 10,  # PopulerMi
        'N': 8,   # YeniMi
        'O': 10,  # OneCikan
        'P': 10,  # CokSatan
        'Q': 12,  # MarkaVitrini
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    wb.save(output_path)
    print(f"Excel dosyası kaydedildi: {output_path}")
    return len(products)

# Ana çalıştırma
if __name__ == "__main__":
    pricelist_path = "/mnt/uploads/ses_3b3d28946ffe2yh5rSr2HGXvow/PriceList_2025_GBP.pdf"
    gp_path = "/mnt/workspace/258iT5oxNEFM9USsyJp4GKR9LsRrdMyu63pcuFxepD6i7PR/GP_ENG_2025.pdf"
    output_path = "/mnt/workspace/258iT5oxNEFM9USsyJp4GKR9LsRrdMyu63pcuFxepD6i7PR/Beta_Urunler_2026.xlsx"
    
    print("Veriler çıkarılıyor...")
    products = extract_all_data(pricelist_path, gp_path, 75)
    print(f"Toplam {len(products)} ürün bulundu")
    
    print("\nExcel oluşturuluyor...")
    count = create_excel(products, output_path)
    print(f"\nTamamlandı! {count} ürün Excel'e yazıldı.")
    
    # İlk 5 ürünü göster
    print("\nÖrnek ürünler:")
    for p in products[:5]:
        print(f"SKU: {p['StokKodu']}, Fiyat: {p['Fiyat']} TRY, Ürün: {p['UrunAdi'][:50]}...")
